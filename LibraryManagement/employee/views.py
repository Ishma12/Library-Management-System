from django.shortcuts import render, redirect,get_object_or_404, reverse
from django.http import HttpResponse
from .models import Book
from .models import BorrowedBook
from django.http import JsonResponse
from datetime import datetime  
from django.contrib.auth import get_user_model
from django.views.decorators.csrf import csrf_exempt  
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.views.generic.base import View
from xhtml2pdf import pisa
from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile
from .models import BookRequest 
from .forms import BookForm, EditBookForm, EditBorrowedBookForm, AddBorrowedBookForm
from library.models import Notification
from copy import deepcopy


User=get_user_model()

def detail(request,book_id):
    book= get_object_or_404(Book,id=book_id)
    return render(request, 'employee/detail.html', {"book":book})



def bookrequest(request):
    if request.method == 'POST':
        book_name = request.POST.get('bookName')
        author = request.POST.get('author')
        category = request.POST.get('category')

        # Save the book request to the database
        BookRequest.objects.create(
            student=request.user,
            book_name=book_name,
            author=author,
            category=category
        )

        return JsonResponse({'message': 'success'})
    book_requests=BookRequest.objects.all().order_by("-id")    
    return render(request, 'employee/requestfromstudent.html', {"book_requests": book_requests})


def approve_bookrequest(request,book_id):
    req= BookRequest.objects.get(id=book_id)
    req.is_approved=True
    req.save()
    Notification.objects.create(detail=f"Your request for {req.book_name} has been approved. You can visit library.", user=req.student)
    return redirect("employee-requestfromstudent")


def decline_bookrequest(request,book_id):
    req= BookRequest.objects.get(id=book_id)
    req.is_approved=False
    req.save()
    Notification.objects.create(detail=f"Your request for {req.book_name} has been declined.", user=req.student)
    return redirect("employee-requestfromstudent") 
   

def generate_borrowed_books_excel(request):
    # Query the BorrowedBook data
    borrowed_books = BorrowedBook.objects.all()

    # Create a new Workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the sheet
    headers = ['ID', 'Book Name', 'Borrowed Date', 'Returned Date', 'Fine']
    ws.append(headers)

    # Add data to the sheet
    for borrowed_book in borrowed_books:
        data = [
                borrowed_book.book.isbn, # Use the 'id' attribute directly
                borrowed_book.book.book_name,
                # borrowed_book.student.name,
                borrowed_book.borrowed_date,
                borrowed_book.returned_date,
                borrowed_book.fine,
        ]
        ws.append(data)

    # Apply date formatting to the date cells in the Borrowed Date and Returned Date columns
    date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')

    # Borrowed Date column (column E)
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row):
        for cell in row:
            cell.style = date_style

    # Returned Date column (column F)
    for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
        for cell in row:
            cell.style = date_style

    # Set column width for date columns
    date_columns = [5, 6]  # Columns E (Borrowed Date) and F (Returned Date)
    for col_num in date_columns:
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15  # Adjust the width as needed

    # Save the workbook to a temporary file
    with NamedTemporaryFile(delete=False) as tmpfile:
        wb.save(tmpfile.name)

    # Create a FileResponse and delete the temporary file after serving
    response = FileResponse(open(tmpfile.name, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'inline; filename=borrowed_books.xlsx'

    return response



class GeneratePDFReportView(View):
    def get(self, request, *args, **kwargs):
        books = Book.objects.all()  
        template_path = 'employee/pdf_report_template.html'  
        context = {'books': books}
        
        # Render the template
        template = get_template(template_path)
        html = template.render(context)

        # Create a PDF response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'filename="book_report.pdf"'
        pisa_status = pisa.CreatePDF(html, dest=response)

        # Return the PDF response if the PDF was created successfully
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response


@login_required
def employee(request):
    books = Book.objects.all()
     # Fetch the count of books and borrowed books
    book_count = Book.objects.count()
    borrowed_book_count = BorrowedBook.objects.count()
    return render(request, 'employee/edashboard.html', {'books': books, 'book_count': book_count, 'borrowed_book_count': borrowed_book_count})

@login_required
def book(request):
    books = Book.objects.all()
    borrowed_books = BorrowedBook.objects.all()
    book_count = Book.objects.count()
    borrowed_book_count = BorrowedBook.objects.count()
    return render(request, 'employee/ebook.html', {'books': books,'book_count': book_count,  'borrowed_books': borrowed_books,'borrowed_book_count': borrowed_book_count})

@login_required
def borrowedbook(request):
    borrowed_books = BorrowedBook.objects.all().order_by('-id')
    book_count = Book.objects.count()
    borrowed_book_count = BorrowedBook.objects.count()
    return render(request, 'employee/eborrowedbook.html', {'book_count': book_count,  'borrowed_books': borrowed_books,'borrowed_book_count': borrowed_book_count})


def changepw(request):
    return render(request, 'employee/changepw.html')




@login_required
def addbook(request):

    if request.method == "POST":
       
        # create a form instance and populate it with data from the request:
        form = BookForm(request.POST, request.FILES)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            form.save()
            # ...
            # redirect to a new URL:
            return redirect(reverse ('employee-ebook'))

    # if a GET (or any other method) we'll create a blank form
    else:
        form = BookForm()
    

    return render(request, 'employee/addbook.html', {"form": form})

@login_required
def editbook(request, book_id):
    book=get_object_or_404(Book,id= book_id)
    form= EditBookForm(instance=book)
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        form = EditBookForm(request.POST, request.FILES,instance=book)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            form.save()
            # ...
            # redirect to a new URL:
            return redirect(reverse ('employee-ebook'))
    return render(request, 'employee/editbook.html', {"form": form})


@login_required
def deletebook(request, book_id):
    book=get_object_or_404(Book,id= book_id)   
    book.delete()
    # If the request method is not POST, render the deletebook.html template
    return redirect(reverse ('employee-ebook'))

@login_required
def addborrowedbook(request):
    if request.method == 'POST':
     if request.method == 'POST':
        form = AddBorrowedBookForm(request.POST)
        if form.is_valid():
            borrowedbook= form.save()
            borrowedbook.is_borrowed= True
            borrowedbook.save()
            # Redirect to a success page or perform other actions
            return redirect(reverse ('employee-eborrowedbook'))
    else:
        form = AddBorrowedBookForm()  


    return render(request, 'employee/addborrowedbook.html', {"form":form})



@login_required
def editborrowedbook(request, borrowbook_id):
    borrowedbook=get_object_or_404(BorrowedBook,id=borrowbook_id)
    form= EditBorrowedBookForm(instance=borrowedbook)

    
    if request.method == 'POST':
        prev=deepcopy(borrowedbook)
        # create a form instance and populate it with data from the request:
        form = EditBorrowedBookForm(request.POST,instance=borrowedbook)
        # check whether it's valid:
      
        if form.is_valid():
            
            new_borrowed_date = form.cleaned_data.get('borrowed_date')
            new_returned_date = form.cleaned_data.get('returned_date')
            new_fine = form.cleaned_data.get('fine')
            if new_fine !=  prev.fine:
                Notification.objects.create(detail=f"You have been fined {new_fine} rupees for {borrowedbook.book.book_name}.", user=borrowedbook.student)
                print("Notified.")
            borrowedbook.borrowed_date = new_borrowed_date
            borrowedbook.returned_date = new_returned_date
            borrowedbook.fine = new_fine
            borrowedbook.save()
           
            return redirect(reverse ('employee-eborrowedbook'))
    return render(request, 'employee/editborrowedbook.html', {"form": form,'bbook':borrowedbook})

  

@login_required
def deleteborrowedbook(request):
    if request.method == 'POST':
        borrowedbook_id = request.POST.get('borrowedbookID')

        # Use filter instead of get to handle multiple borrowed books with the same ID
        borrowed_books = BorrowedBook.objects.filter(borrowedbook_id=borrowedbook_id)

        if borrowed_books.exists():
            # Iterate over the queryset and delete each borrowed book
            for borrowed_book in borrowed_books:
                borrowed_book.delete()

            # Return a success message as JSON response
            return JsonResponse({'message': 'Borrowed books deleted successfully'})
        else:
            # Return an error message as JSON response if no borrowed books are found
            return JsonResponse({'error': 'No borrowed books found for the specified ID'}, status=400)

    # If the request method is not POST, render the deleteborrowedbook.html template
    return render(request, 'employee/deleteborrowedbook.html')

#student ko view
def borrowbook(request,book_id):
    if request.user.usertype == User.EMPLOYEE:
        return redirect(reverse('book-detail', kwargs={"book_id": book_id}))

    if request.method =='POST':
        book=get_object_or_404(Book, id=book_id)
        borrowed_date = request.POST.get('borrowedDate')
        returned_date = request.POST.get('returnedDate')

        BorrowedBook.objects.create(
            borrowed_date=borrowed_date,
            returned_date=returned_date,
            student=request.user,
            book=book

        )

    return redirect(reverse('student-detailbook', kwargs={"book_id": book_id}))


def approve_borrowbook(request, borrowbook_id):
    borrowedbook=get_object_or_404(BorrowedBook,id=borrowbook_id)
    borrowedbook.is_borrowed=True
    borrowedbook.save()
    Notification.objects.create(detail=f"Your book borrow request for {borrowedbook.book.book_name} has been approved. You can visit library.", user=borrowedbook.student)
    return redirect(reverse('employee-eborrowedbook'))

def decline_borrowbook(request, borrowbook_id):
    borrowedbook=get_object_or_404(BorrowedBook,id=borrowbook_id)
    borrowedbook.is_borrowed=False
    borrowedbook.save()
    Notification.objects.create(detail=f"Your book borrow request for {borrowedbook.book.book_name} has been declined.", user=borrowedbook.student)
    return redirect(reverse('employee-eborrowedbook'))




